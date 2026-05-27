import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 文件路径
input_dir = r"D:\Joker\Timoshenko\data\2d_ssss"
output_file = r"D:\Joker\Timoshenko\data\2d_ssss\2d_ssss_vibration_q4_comparison.xlsx"

# 文件映射: (输入文件名, Sheet名称)
files = [
    ("2d_ssss_vibration_q4_17x17_mindlin_exact_comparison.csv", "17x17"),
    ("2d_ssss_vibration_q4_5x5_mindlin_exact_comparison.csv", "5x5"),
]

# 列映射：新列名 -> CSV中的列名
columns_map = {
    "Model": None,  # 手动填充
    "m": "m",
    "n": "n",
    "\\Omega_exact": "Omega_exact",
    "\\Omega_FEM": "Omega_num",
    "誤差": "rel_error_percent",
}

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    for csv_file, sheet_name in files:
        csv_path = f"{input_dir}\\{csv_file}"
        df = pd.read_csv(csv_path)

        # 创建新的DataFrame
        new_df = pd.DataFrame()
        for new_col, src_col in columns_map.items():
            if src_col is None:
                # 从文件名中提取网格信息作为Model
                if "17x17" in csv_file:
                    new_df[new_col] = ["Q4-17×17"] * len(df)
                else:
                    new_df[new_col] = ["Q4-5×5"] * len(df)
            else:
                new_df[new_col] = df[src_col]

        # 写入Excel
        new_df.to_excel(writer, sheet_name=sheet_name, index=False)

# 调整列宽
wb = load_workbook(output_file)
for ws in wb.worksheets:
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col_cells:
            if cell.value is not None:
                # 处理中文字符宽度
                val_str = str(cell.value)
                char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                max_len = max(max_len, char_len)
        ws.column_dimensions[col_letter].width = max_len + 2

wb.save(output_file)
print(f"Excel 文件已生成: {output_file}")
print("包含 Sheet: 17x17, 5x5")
