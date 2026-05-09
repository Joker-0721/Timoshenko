from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
import html
import math

from openpyxl import load_workbook, Workbook
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "results"
OUT = ROOT / "reports" / "bui_style"


@dataclass(frozen=True)
class Series:
    label: str
    x: list[float]
    y: list[float]
    color: str = "#111111"
    dash: str | None = None
    marker: str = "circle"


def read_rows(filename: str) -> list[dict[str, object]]:
    wb = load_workbook(RESULTS / filename, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows: list[dict[str, object]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def read_sheet_rows(filename: str, sheet_name: str) -> tuple[list[str], list[list[object]]]:
    wb = load_workbook(RESULTS / filename, data_only=True)
    ws = wb[sheet_name]
    headers = [str(cell.value) for cell in ws[1]]
    rows: list[list[object]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in values):
            continue
        rows.append(list(values))
    return headers, rows


def pct_error(row: dict[str, object]) -> float:
    return float(row["k_error"]) * 100.0


def fmt(value: object, digits: int = 4) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if abs(value) >= 1000:
            return f"{value:.2f}"
        return f"{value:.{digits}f}".rstrip("0").rstrip(".")
    return str(value)


def write_png(path: Path, title: str, xlabel: str, ylabel: str, series: list[Series]) -> None:
    width, height = 1200, 780
    left, right, top, bottom = 140, 70, 92, 110
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype(r"C:\Windows\Fonts\arial.ttf", 17)
        small_font = ImageFont.truetype(r"C:\Windows\Fonts\arial.ttf", 15)
        title_font = ImageFont.truetype(r"C:\Windows\Fonts\arialbd.ttf", 20)
    except OSError:
        font = ImageFont.load_default()
        small_font = ImageFont.load_default()
        title_font = ImageFont.load_default()

    xs = [v for s in series for v in s.x]
    ys = [v for s in series for v in s.y]
    xmin, xmax = min(xs), max(xs)
    ymin, ymax = min(ys), max(ys)
    if xmin == xmax:
        xmin -= 1
        xmax += 1
    ypad = (ymax - ymin) * 0.08 if ymax > ymin else 1
    ymin = max(0, ymin - ypad)
    ymax = ymax + ypad

    def px(x: float) -> float:
        return left + (x - xmin) / (xmax - xmin) * (width - left - right)

    def py(y: float) -> float:
        return height - bottom - (y - ymin) / (ymax - ymin) * (height - top - bottom)

    # Axes and light horizontal grid.
    draw.line([(left, top), (left, height - bottom), (width - right, height - bottom)], fill="#111111", width=2)
    for i in range(6):
        y = ymin + i * (ymax - ymin) / 5
        yy = py(y)
        draw.line([(left, yy), (width - right, yy)], fill="#DDDDDD", width=1)
        draw.text((left - 84, yy - 9), fmt(y, 2), fill="#111111", font=small_font)
    for i in range(6):
        x = xmin + i * (xmax - xmin) / 5
        xx = px(x)
        draw.line([(xx, height - bottom), (xx, height - bottom + 7)], fill="#111111", width=1)
        draw.text((xx - 18, height - bottom + 15), fmt(x, 2), fill="#111111", font=small_font)

    draw.text((left, 35), title, fill="#111111", font=title_font)
    xbbox = draw.textbbox((0, 0), xlabel, font=font)
    draw.text(((width - (xbbox[2] - xbbox[0])) / 2, height - 42), xlabel, fill="#111111", font=font)
    ylabel_img = Image.new("RGBA", (360, 32), (255, 255, 255, 0))
    ylabel_draw = ImageDraw.Draw(ylabel_img)
    ylabel_draw.text((0, 2), ylabel, fill="#111111", font=font)
    ylabel_img = ylabel_img.rotate(90, expand=True)
    img.paste(ylabel_img, (24, top + (height - top - bottom - ylabel_img.height) // 2), ylabel_img)

    legend_x, legend_y = width - right - 240, top + 10
    for idx, s in enumerate(series):
        pts = [(px(x), py(y)) for x, y in zip(s.x, s.y)]
        color = s.color
        if s.dash:
            for a, b in zip(pts, pts[1:]):
                dashed_line(draw, a, b, color)
        else:
            draw.line(pts, fill=color, width=3)
        for x, y in pts:
            if s.marker == "square":
                draw.rectangle((x - 5, y - 5, x + 5, y + 5), outline=color, fill="white", width=2)
            else:
                draw.ellipse((x - 5, y - 5, x + 5, y + 5), outline=color, fill="white", width=2)
        ly = legend_y + idx * 28
        if s.dash:
            dashed_line(draw, (legend_x, ly + 9), (legend_x + 36, ly + 9), color)
        else:
            draw.line([(legend_x, ly + 9), (legend_x + 36, ly + 9)], fill=color, width=3)
        draw.text((legend_x + 45, ly), s.label, fill="#111111", font=small_font)

    img.save(path)


def dashed_line(draw: ImageDraw.ImageDraw, a: tuple[float, float], b: tuple[float, float], color: str) -> None:
    ax, ay = a
    bx, by = b
    length = math.hypot(bx - ax, by - ay)
    if length == 0:
        return
    dash, gap = 12, 8
    steps = int(length // (dash + gap)) + 1
    ux, uy = (bx - ax) / length, (by - ay) / length
    for i in range(steps):
        start = i * (dash + gap)
        end = min(start + dash, length)
        if start >= length:
            break
        draw.line(
            [(ax + ux * start, ay + uy * start), (ax + ux * end, ay + uy * end)],
            fill=color,
            width=3,
        )


def write_svg(path: Path, title: str, xlabel: str, ylabel: str, series: list[Series]) -> None:
    width, height = 720, 470
    left, right, top, bottom = 76, 36, 54, 68
    xs = [v for s in series for v in s.x]
    ys = [v for s in series for v in s.y]
    xmin, xmax = min(xs), max(xs)
    ymin, ymax = min(ys), max(ys)
    if xmin == xmax:
        xmin -= 1
        xmax += 1
    ypad = (ymax - ymin) * 0.08 if ymax > ymin else 1
    ymin = max(0, ymin - ypad)
    ymax = ymax + ypad

    def px(x: float) -> float:
        return left + (x - xmin) / (xmax - xmin) * (width - left - right)

    def py(y: float) -> float:
        return height - bottom - (y - ymin) / (ymax - ymin) * (height - top - bottom)

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="white"/>',
        '<style>text{font-family:Arial,Helvetica,sans-serif;font-size:13px;fill:#111}.title{font-size:16px;font-weight:700}.axis{stroke:#111;stroke-width:1.4}.grid{stroke:#ddd;stroke-width:1}</style>',
        f'<text class="title" x="{left}" y="28">{html.escape(title)}</text>',
    ]
    for i in range(6):
        y = ymin + i * (ymax - ymin) / 5
        yy = py(y)
        parts.append(f'<line class="grid" x1="{left}" y1="{yy:.1f}" x2="{width-right}" y2="{yy:.1f}"/>')
        parts.append(f'<text x="{left-10}" y="{yy+4:.1f}" text-anchor="end">{fmt(y, 2)}</text>')
    parts.append(f'<line class="axis" x1="{left}" y1="{top}" x2="{left}" y2="{height-bottom}"/>')
    parts.append(f'<line class="axis" x1="{left}" y1="{height-bottom}" x2="{width-right}" y2="{height-bottom}"/>')
    for i in range(6):
        x = xmin + i * (xmax - xmin) / 5
        xx = px(x)
        parts.append(f'<line class="axis" x1="{xx:.1f}" y1="{height-bottom}" x2="{xx:.1f}" y2="{height-bottom+6}"/>')
        parts.append(f'<text x="{xx:.1f}" y="{height-bottom+24}" text-anchor="middle">{fmt(x, 2)}</text>')
    parts.append(f'<text x="{width/2:.1f}" y="{height-18}" text-anchor="middle">{html.escape(xlabel)}</text>')
    parts.append(f'<text x="18" y="{top+8}" transform="rotate(-90 18,{top+8})" text-anchor="end">{html.escape(ylabel)}</text>')

    legend_x, legend_y = width - right - 180, top + 10
    for idx, s in enumerate(series):
        points = " ".join(f"{px(x):.1f},{py(y):.1f}" for x, y in zip(s.x, s.y))
        dash = ' stroke-dasharray="8 6"' if s.dash else ""
        parts.append(f'<polyline points="{points}" fill="none" stroke="{s.color}" stroke-width="2.2"{dash}/>')
        for x, y in zip(s.x, s.y):
            xx, yy = px(x), py(y)
            if s.marker == "square":
                parts.append(f'<rect x="{xx-4:.1f}" y="{yy-4:.1f}" width="8" height="8" fill="white" stroke="{s.color}" stroke-width="1.8"/>')
            else:
                parts.append(f'<circle cx="{xx:.1f}" cy="{yy:.1f}" r="4" fill="white" stroke="{s.color}" stroke-width="1.8"/>')
        ly = legend_y + idx * 22
        parts.append(f'<line x1="{legend_x}" y1="{ly}" x2="{legend_x+28}" y2="{ly}" stroke="{s.color}" stroke-width="2.2"{dash}/>')
        parts.append(f'<text x="{legend_x+36}" y="{ly+4}">{html.escape(s.label)}</text>')
    parts.append("</svg>")
    path.write_text("\n".join(parts), encoding="utf-8")


def md_table(headers: list[str], rows: Iterable[list[object]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(fmt(v) for v in row) + " |")
    return "\n".join(lines)


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[object]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.style = "Headline 3"
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max(len(fmt(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 24)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    uniaxial = read_rows("mindlin_uniaxial_ssss_table_9_1.xlsx")
    shear = read_rows("mindlin_shear_ssss_table_9_12.xlsx")
    combined = read_rows("mindlin_combined_load_ssss_table_9_13.xlsx")
    biaxial = read_rows("mindlin_biaxial_ssss.xlsx")
    bui_workbook = RESULTS / "bui_2011_table1_table2.xlsx"
    bui_table1_headers: list[str] = []
    bui_table1_rows: list[list[object]] = []
    bui_table2_headers: list[str] = []
    bui_table2_rows: list[list[object]] = []
    if bui_workbook.exists():
        bui_table1_headers, bui_table1_rows = read_sheet_rows(bui_workbook.name, "Table 1")
        bui_table2_headers, bui_table2_rows = read_sheet_rows(bui_workbook.name, "Table 2")

    datasets = {
        "fig_uniaxial": (
            "Variation of uniaxial buckling load factor",
            "Aspect ratio a/b",
            "Buckling load factor k",
            [
                Series("Present", [float(r["a/b"]) for r in uniaxial], [float(r["k_num"]) for r in uniaxial], "#111111"),
                Series("Exact", [float(r["a/b"]) for r in uniaxial], [float(r["k_exact"]) for r in uniaxial], "#666666", "dash", "square"),
            ],
        ),
        "fig_shear": (
            "Variation of shear buckling load factor",
            "Aspect ratio a/b",
            "Shear buckling load factor k",
            [
                Series("Present", [float(r["a/b"]) for r in shear], [float(r["k_num"]) for r in shear], "#111111"),
                Series("Reference", [float(r["a/b"]) for r in shear], [float(r["k_book"]) for r in shear], "#666666", "dash", "square"),
            ],
        ),
        "fig_biaxial": (
            "Variation of biaxial buckling load factor",
            "Aspect ratio gamma = Ny/Nx",
            "Biaxial buckling load factor k",
            [
                Series("Present", [float(r["gamma_Ny_over_Nx"]) for r in biaxial], [float(r["k_num"]) for r in biaxial], "#111111"),
                Series("Exact", [float(r["gamma_Ny_over_Nx"]) for r in biaxial], [float(r["k_exact"]) for r in biaxial], "#666666", "dash", "square"),
            ],
        ),
        "fig_combined": (
            "Variation of buckling factor under combined loads",
            "Aspect ratio delta = sigma/tau",
            "Combined buckling load factor k",
            [
                Series("Present", [float(r["sigma_over_tau"]) for r in combined], [float(r["k_num"]) for r in combined], "#111111"),
                Series("Reference", [float(r["sigma_over_tau"]) for r in combined], [float(r["k_book"]) for r in combined], "#666666", "dash", "square"),
            ],
        ),
    }

    for stem, (title, xlabel, ylabel, series) in datasets.items():
        write_svg(OUT / f"{stem}.svg", title, xlabel, ylabel, series)
        write_png(OUT / f"{stem}.png", title, xlabel, ylabel, series)

    uniaxial_rows = [[r["a/b"], r["nodes"], r["k_exact"], r["k_num"], pct_error(r)] for r in uniaxial]
    shear_rows = [[r["a/b"], r["nodes"], r["k_book"], r["k_num"], pct_error(r)] for r in shear]
    biaxial_rows = [[r["gamma_Ny_over_Nx"], r["nodes"], r["k_exact"], r["k_num"], pct_error(r)] for r in biaxial]
    combined_rows = [[r["sigma_over_tau"], r["nodes"], r["k_book"], r["k_num"], pct_error(r)] for r in combined]

    report = f"""# Bui-style Buckling Tables and Figures

Generated from `Timoshenko/results`.

The layout follows the comparison style used in Bui et al. (2011): reference values beside the present numerical values, followed by a compact relative error column and line figures showing the variation of normalized buckling factor `k`.

## Fig. 1. Uniaxial Compression, SSSS Rectangular Plates

![Uniaxial buckling factor](fig_uniaxial.png)

{md_table(["a/b", "nodes", "Exact k", "Present k", "Error (%)"], uniaxial_rows)}

## Fig. 2. Pure Shear, SSSS Rectangular Plates

![Shear buckling factor](fig_shear.png)

{md_table(["a/b", "nodes", "Reference k", "Present k", "Error (%)"], shear_rows)}

## Fig. 3. Biaxial Compression, SSSS Square Plate

![Biaxial buckling factor](fig_biaxial.png)

{md_table(["gamma = Ny/Nx", "nodes", "Exact k", "Present k", "Error (%)"], biaxial_rows)}

## Fig. 4. Combined Compression and Shear, SSSS Square Plate

![Combined-load buckling factor](fig_combined.png)

{md_table(["sigma/tau", "nodes", "Reference k", "Present k", "Error (%)"], combined_rows)}

## Bui et al. (2011) Table 1. Uniaxial Buckling, Square Plate

Square plate, `a/b = 1`, `t/b = 0.01`, uniaxial compression along the x-axis. Boundary labels follow Bui et al.; the recomputed columns use the weak form in `md/test2.md`.

{md_table(bui_table1_headers, bui_table1_rows) if bui_table1_rows else "_Run `julia --project=. src/generate_bui_2011_table1_table2.jl` to generate this table._"}

## Bui et al. (2011) Table 2. Critical Uniaxial Buckling Factor Comparison

The published reference columns are copied from the PDF table; only `Present` is recomputed from the local formula and the 13 x 13 case.

{md_table(bui_table2_headers, bui_table2_rows) if bui_table2_rows else "_Run `julia --project=. src/generate_bui_2011_table1_table2.jl` to generate this table._"}

## Notes

- `Error (%) = abs(Present - reference) / reference * 100`.
- These figures deliberately use the same plain academic style as the paper: white background, black primary line, dashed reference line, and minimal decoration.
- The present values are consistently above the reference values. This suggests a systematic normalization or geometric-stiffness scaling issue rather than only a plotting issue.
"""
    (OUT / "bui_style_summary.md").write_text(report, encoding="utf-8")

    latex = "\n\n".join(
        [
            latex_table("Uniaxial compression, SSSS rectangular plates", ["a/b", "nodes", "Exact $k$", "Present $k$", "Error (\\%)"], uniaxial_rows),
            latex_table("Pure shear, SSSS rectangular plates", ["a/b", "nodes", "Reference $k$", "Present $k$", "Error (\\%)"], shear_rows),
            latex_table("Biaxial compression, SSSS square plate", ["$\\gamma=N_y/N_x$", "nodes", "Exact $k$", "Present $k$", "Error (\\%)"], biaxial_rows),
            latex_table("Combined compression and shear, SSSS square plate", ["$\\sigma/\\tau$", "nodes", "Reference $k$", "Present $k$", "Error (\\%)"], combined_rows),
            latex_table("Bui et al. (2011) Table 1: uniaxial buckling load factor of a square plate", bui_table1_headers, bui_table1_rows) if bui_table1_rows else "",
            latex_table("Bui et al. (2011) Table 2: critical uniaxial buckling load factor comparison", bui_table2_headers, bui_table2_rows) if bui_table2_rows else "",
        ]
    )
    (OUT / "tables.tex").write_text(latex, encoding="utf-8")

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "Uniaxial", ["a/b", "nodes", "Exact k", "Present k", "Error (%)"], uniaxial_rows)
    add_sheet(wb, "Shear", ["a/b", "nodes", "Reference k", "Present k", "Error (%)"], shear_rows)
    add_sheet(wb, "Biaxial", ["gamma = Ny/Nx", "nodes", "Exact k", "Present k", "Error (%)"], biaxial_rows)
    add_sheet(wb, "Combined", ["sigma/tau", "nodes", "Reference k", "Present k", "Error (%)"], combined_rows)
    if bui_table1_rows:
        add_sheet(wb, "Bui 2011 Table 1", bui_table1_headers, bui_table1_rows)
    if bui_table2_rows:
        add_sheet(wb, "Bui 2011 Table 2", bui_table2_headers, bui_table2_rows)
    wb.save(OUT / "bui_style_tables.xlsx")


def latex_table(caption: str, headers: list[str], rows: list[list[object]]) -> str:
    cols = "c" * len(headers)
    body = [
        "\\begin{table}[htbp]",
        "\\centering",
        f"\\caption{{{caption}}}",
        f"\\begin{{tabular}}{{{cols}}}",
        "\\hline",
        " & ".join(headers) + r" \\",
        "\\hline",
    ]
    for row in rows:
        body.append(" & ".join(fmt(v) for v in row) + r" \\")
    body.extend(["\\hline", "\\end{tabular}", "\\end{table}"])
    return "\n".join(body)


if __name__ == "__main__":
    main()
